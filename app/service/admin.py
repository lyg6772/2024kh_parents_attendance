from datetime import datetime, timedelta
from app.dao.admin import AdminAttendeeDao
from fastapi import Depends, Request
from dateutil.relativedelta import relativedelta
from fastapi.templating import Jinja2Templates
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO


templates = Jinja2Templates(directory="./app/template")


class AdminAttendeeService:
    def __init__(self, dao: AdminAttendeeDao = Depends(AdminAttendeeDao)):
        self.dao = dao

    async def get_attendee_table(self, request: Request, date_str):
        start_dt = datetime.strptime(f'{date_str}01', '%Y%m%d')
        end_dt = start_dt + relativedelta(months=1) - timedelta(days=1)
        before_month_str = datetime.strftime(start_dt - relativedelta(months=1), '%Y%m')
        after_month_str = datetime.strftime(start_dt + relativedelta(months=1), '%Y%m')
        attendance_raw = await self.dao.get_attendee(
            start_dt=start_dt.strftime('%Y%m%d'), end_dt=end_dt.strftime('%Y%m%d')
        )
        attendee_dict = dict()
        for attendance in attendance_raw:
            # 컬럼명을 소문자로 변환하여 접근
            atdc_date = attendance.get('atdc_date') or attendance.get('ATDC_DATE')
            atde_name = attendance.get('atde_name') or attendance.get('ATDE_NAME')
            if atdc_date:
                attendee_dict[atdc_date] = atde_name or ''

        notice_raw = await self.dao.get_notice(
            start_dt=start_dt.strftime('%Y%m%d'), end_dt=end_dt.strftime('%Y%m%d')
        )
        notice_dict = dict()
        for notice in notice_raw:
            # 컬럼명을 소문자로 변환하여 접근
            atdc_date = notice.get('atdc_date') or notice.get('ATDC_DATE')
            atdc_notice = notice.get('atdc_notice') or notice.get('ATDC_NOTICE')
            if atdc_date:
                notice_dict[atdc_date] = atdc_notice or ''

        starting_weekday = start_dt.isoweekday()
        num_days = (end_dt - start_dt).days + 1
        calendar = []
        week = []
        if starting_weekday != 7:
            for i in range(starting_weekday):
                week.append({"day": None, "attendee": '', 'notice': '', "date": None})
        for day in range(1, num_days + 1):
            cal_date = f"{date_str}{day:02d}"
            week.append(
                {"day": day, "attendee": attendee_dict.get(cal_date, ''), 'notice': notice_dict.get(cal_date, ''),
                 "date": cal_date})
            if len(week) == 7:
                calendar.append(week)
                week = []
        if week:
            while len(week) < 7:
                week.append({"day": None, "attendee": '', 'notice': '', "date": None})
            calendar.append(week)

        return templates.TemplateResponse('./admin_attendee.html', context={
            "year": start_dt.year,
            "month": start_dt.month,
            "cal_date_formatted": date_str,
            "prev_month": before_month_str,
            "next_month": after_month_str,
            "calendar": calendar,
            "request": request
        })

    async def post_attendee(self, request:Request):
        json_data = await request.json()
        attendee_list = json_data.get('attendee', '').split(',')
        notice = json_data.get('notice', '')
        attendee_date = json_data.get('date')
        await self.dao.insert_attendee(attendee_date=attendee_date, attendee_list=attendee_list)
        await self.dao.insert_notice(attendee_date=attendee_date, notice=notice)

        return Response(status_code=200, content="OK")

    async def export_to_excel(self, date_str):
        """캘린더 데이터를 Excel로 export"""
        start_dt = datetime.strptime(f'{date_str}01', '%Y%m%d')
        end_dt = start_dt + relativedelta(months=1) - timedelta(days=1)

        # DB에서 데이터 조회
        attendance_raw = await self.dao.get_attendee(
            start_dt=start_dt.strftime('%Y%m%d'), end_dt=end_dt.strftime('%Y%m%d')
        )
        attendee_dict = dict()
        for attendance in attendance_raw:
            atdc_date = attendance.get('atdc_date') or attendance.get('ATDC_DATE')
            atde_name = attendance.get('atde_name') or attendance.get('ATDE_NAME')
            if atdc_date:
                attendee_dict[atdc_date] = atde_name or ''

        notice_raw = await self.dao.get_notice(
            start_dt=start_dt.strftime('%Y%m%d'), end_dt=end_dt.strftime('%Y%m%d')
        )
        notice_dict = dict()
        for notice in notice_raw:
            atdc_date = notice.get('atdc_date') or notice.get('ATDC_DATE')
            atdc_notice = notice.get('atdc_notice') or notice.get('ATDC_NOTICE')
            if atdc_date:
                notice_dict[atdc_date] = atdc_notice or ''

        # Excel 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = f'{start_dt.year}년 {start_dt.month}월'

        # 스타일 정의
        header_fill = PatternFill(start_color='4F63F7', end_color='4F63F7', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 제목
        ws.merge_cells('A1:C1')
        title_cell = ws['A1']
        title_cell.value = f'{start_dt.year}년 {start_dt.month}월 보람교사 현황'
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # 헤더
        headers = ['날짜', '참석자', '특이사항']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        # 데이터 입력
        row = 4
        num_days = (end_dt - start_dt).days + 1
        for day in range(1, num_days + 1):
            cal_date = f"{date_str}{day:02d}"

            # 날짜 포맷 (2026-03-01)
            date_obj = datetime.strptime(cal_date, '%Y%m%d')
            date_str_formatted = date_obj.strftime('%Y-%m-%d')

            # 날짜 셀
            date_cell = ws.cell(row=row, column=1)
            date_cell.value = date_str_formatted
            date_cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
            date_cell.border = border

            # 참석자 셀
            attendee_cell = ws.cell(row=row, column=2)
            attendee_cell.value = attendee_dict.get(cal_date, '')
            attendee_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            attendee_cell.border = border

            # 특이사항 셀
            notice_cell = ws.cell(row=row, column=3)
            notice_cell.value = notice_dict.get(cal_date, '')
            notice_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            notice_cell.border = border

            row += 1

        # 컬럼 너비 설정
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40

        # 행 높이 설정
        ws.row_dimensions[3].height = 25
        for r in range(4, row):
            ws.row_dimensions[r].height = 30

        # BytesIO에 저장
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return excel_file
